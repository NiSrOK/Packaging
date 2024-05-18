import random
from genetic2D import genetic
from openpyxl import Workbook, load_workbook
import timeit
from exhaustive_search2D import exhaustive_search

def generate_base_population(n):
    list_of_segments = []
    for i in range(n):
        a = random.randint(1,20)
        b = random.randint(1,20)
        p = random.randint(1,20)
        name = i
        list_of_segments.append((a, b, p, name))
    return list_of_segments

def tuner(c_x, c_y, h1, h2):
    count_columns = 3
    wb = Workbook()
    ws = wb.active
    ws[f'A1'] = 'Отклонение генетический 1 набор'
    ws[f'B1'] = 'Отклонение генетический 20 наборов'
    ws[f'C1'] = 'Отклонение ПП лучший'
    ws[f'D1'] = 'Отклонение ПП худший'
    ws[f'E1'] = 'Время генетический 1 набор'
    ws[f'F1'] = 'Время генетический 20 наборов'
    ws[f'G1'] = 'Время ПП'
    row = 2
    for _ in range(20):
        population = generate_base_population(10)
        list_of_segments = population

        start_time_genetic = timeit.default_timer()
        gen_dev_total_1 = genetic(list_of_segments, 30, 30, c_x, c_y, h1, h2, count_columns, type='calculate')
        end_time_genetic = timeit.default_timer()
        total_time_genetic_1 = end_time_genetic - start_time_genetic
        print(f'Отклонение генетический 1 {gen_dev_total_1}')

        gen_dev_total_more = None
        start_time_genetic = timeit.default_timer()
        for _ in range(20):
            gen_dev = genetic(list_of_segments, 30, 30, c_x, c_y, h1, h2, count_columns, type='calculate')
            if gen_dev_total_more is None or gen_dev_total_more > gen_dev:
                gen_dev_total_more = gen_dev
        end_time_genetic = timeit.default_timer()
        total_time_genetic_more = end_time_genetic - start_time_genetic
        print(f'Отклонение генетический 20 {gen_dev_total_more}')
        
        start_time_exhaustive = timeit.default_timer()
        ex_dev_best, ex_dev_worst = exhaustive_search(list_of_segments, c_x, c_y, h1, h2, count_columns, type='calculate')
        end_time_exhaustive = timeit.default_timer()
        total_time_exhaustive = end_time_exhaustive - start_time_exhaustive
        print(f'Отклонение ПП лучший {ex_dev_best}')
        print(f'Отклонение ПП худший {ex_dev_worst}')
        print('----------------')

        res_gen_1 = float('{:.3f}'.format(gen_dev_total_1))
        res_gen_more = float('{:.3f}'.format(gen_dev_total_more))
        res_exh_best = float('{:.3f}'.format(ex_dev_best))
        res_exh_worst = float('{:.3f}'.format(ex_dev_worst))

        ws[f'A{row}'] = res_gen_1
        ws[f'B{row}'] = res_gen_more
        ws[f'C{row}'] = res_exh_best
        ws[f'D{row}'] = res_exh_worst
        ws[f'E{row}'] = total_time_genetic_1
        ws[f'F{row}'] = total_time_genetic_more
        ws[f'G{row}'] = total_time_exhaustive
        row += 1
        wb.save("tunerVsAfterUpdateY2.xlsx")


c_x = 70 # целевой центр тяжести по x
c_y = 0.5 # целевой центр тяжести по y
h1 = 1 # минимальная длина
# h2 = 120 # максимальная длина
h2 = 80 # максимальная длина
res = tuner(c_x, c_y, h1, h2)
