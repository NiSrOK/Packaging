import random
from genetic2D import genetic
from openpyxl import Workbook, load_workbook
import plotly.graph_objects as go
import timeit
from exhaustive_json import save_exhaustive, read_json

def generate_base_population(n):
    list_of_segments = []
    for i in range(n):
        a = random.randint(1,20)
        b = random.randint(1,20)
        p = random.randint(1,20)
        name = i
        list_of_segments.append((a, b, p, name))
    return list_of_segments

def calculate_length(list_of_segments):
    length = 0
    for i in range(len(list_of_segments)):
        length += list_of_segments[i][0] + list_of_segments[i][1]
    return length

def tuner(c_x, c_y, h1, h2):
    wb = Workbook()
    ws = wb.active
    ws[f'A1'] = 'k'
    ws[f'B1'] = 'g'
    ws[f'C1'] = 'Отклонение от целевого центра тяжести генетический'
    ws[f'D1'] = 'Время расчета генетическим алгоритмом'
    ws[f'E1'] = 'Отклонение от ПП'
    row = 2
    glob_res = []
    exhaustive_res = save_exhaustive(c_x, c_y, h1, h2, 'data.json', 9, 9)
    print(f'Лучшее отклонение ПП: {exhaustive_res[0].best_dev}')
    print(f'Худшее отклонение ПП: {exhaustive_res[0].worst_dev}')
    # exhaustive_res = read_json('data.json')
    # population = generate_base_population(150)
    for k in range(1, 30):
        for g in range(1, 30):
            # res = []
            # count = 0
            # total_time_genetic = 0
            # total_time_exhaustive = 0
            list_of_segments = exhaustive_res[0].list_of_segments

            print(f'k = {k}, g = {g}')
            start_time_genetic = timeit.default_timer()
            gen_dev = genetic(list_of_segments, k, g, c_x, c_y, h1, h2, columns = 1, calculate=True)
            end_time_genetic = timeit.default_timer()
            total_time_genetic = end_time_genetic - start_time_genetic
            
            # start_time_exhaustive = timeit.default_timer()
            # ex_dev = exhaustive_search(list_of_segments, c_x, c_y, h1, h2, calculate=True)
            # end_time_exhaustive = timeit.default_timer()
            # total_time_exhaustive += (end_time_exhaustive - start_time_exhaustive)

            res = (abs(gen_dev-exhaustive_res[0].best_dev))
            res = float('{:.3f}'.format(res))
            # res.append(gen_dev)
            # dev = float('{:.3f}'.format(sum(res) / len(res)))
            dev = float('{:.3f}'.format(gen_dev))
            print(dev)
            if dev == 0:
                dev = 0.001
            glob_res.append((int(k), int(g), dev))
            ws[f'A{row}'] = k
            ws[f'B{row}'] = g
            ws[f'C{row}'] = dev
            ws[f'D{row}'] = total_time_genetic
            ws[f'E{row}'] = res
            row += 1
            wb.save("tuner666.xlsx")
    # wb.save("tuner5.xlsx")
    return glob_res

def plot_3d_graph(data):
    # Разделение данных на отдельные списки для каждой координаты
    x_data, y_data, z_data = zip(*data)

    # Создание трехмерного графика
    fig = go.Figure(data=[go.Scatter3d(x=x_data, y=y_data, z=z_data, mode='markers', marker=dict(color='red', size=3))])

    # Настройка меток осей и установка логарифмической шкалы для оси Z
    fig.update_layout(scene=dict(xaxis_title='k', yaxis_title='g', zaxis_title='Отклонение от целевого центра тяжести', zaxis=dict(type='log')))

    # Отображение графика
    fig.show()

def read_data(file_path):
    # Открываем Excel-файл
    wb = load_workbook(file_path)
    ws = wb.active
    data = []
    # Проходим по строкам и считываем данные
    for row in range(2, ws.max_row + 1):
        # Получаем значения из столбцов A, B, и C для текущей строки
        a_value = int(ws[f'A{row}'].value)
        b_value = int(ws[f'B{row}'].value)
        c_value = float(ws[f'C{row}'].value)

        data.append((a_value, b_value, c_value))

    # Закрываем Excel-файл
    wb.close()

    return data


c_x = 100
c_y = 1.5
h1 = 1 # минимальная длина
h2 = 120 # максимальная длина
res = tuner(c_x, c_y, h1, h2)
# res = read_data('tuner5.xlsx')
# data_points = [(1, 2, 3), (4, 5, 6), (7, 8, 9)]
plot_3d_graph(res)
