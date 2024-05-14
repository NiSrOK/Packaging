import random
from genetic2D import genetic
from openpyxl import Workbook, load_workbook
import timeit
from exhaustive_search2D import exhaustive_search

def generate_base_population(n):
    list_of_segments = []
    for i in range(n):
        a = random.randint(1,20)
        b = random.randint(1,20)
        p = random.randint(1,20)
        name = i
        list_of_segments.append((a, b, p, name))
    return list_of_segments

def tuner(c_x, c_y, h1, h2):
    wb = Workbook()
    ws = wb.active
    ws[f'A1'] = 'Размер списка'
    ws[f'B1'] = 'Отклонение ПП лучший'
    ws[f'C1'] = 'Отклонение ПП худший'
    ws[f'D1'] = 'Время ПП'
    # ws[f'D1'] = 'Время генетический'
    # ws[f'E1'] = 'Время ПП'
    row = 2
    for i in range(12,30):
        population = generate_base_population(i)
        list_of_segments = population

        # start_time_genetic = timeit.default_timer()
        # gen_dev = genetic(list_of_segments, 30, 30, c_x, c_y, h1, h2, columns = 1, calculate=True)
        # end_time_genetic = timeit.default_timer()
        # total_time_genetic = end_time_genetic - start_time_genetic
        # print(f'Отклонение генетический {gen_dev}')
        
        start_time_exhaustive = timeit.default_timer()
        ex_dev_best, ex_dev_worst = exhaustive_search(list_of_segments, c_x, c_y, h1, h2, type='calculate')
        end_time_exhaustive = timeit.default_timer()
        total_time_exhaustive = end_time_exhaustive - start_time_exhaustive
        
        print(f'Размер списка {i}')
        print(f'Отклонение ПП лучший {ex_dev_best}')
        print(f'Отклонение ПП худший {ex_dev_worst}')
        print(f'Время ПП {total_time_exhaustive}')
        print('----------------')

        # res_gen = float('{:.3f}'.format(gen_dev))
        res_exh_best = float('{:.3f}'.format(ex_dev_best))
        res_exh_worst = float('{:.3f}'.format(ex_dev_worst))

        ws[f'A{row}'] = i
        ws[f'B{row}'] = res_exh_best
        ws[f'C{row}'] = res_exh_worst
        ws[f'D{row}'] = total_time_exhaustive
        # ws[f'D{row}'] = total_time_genetic
        # ws[f'E{row}'] = total_time_exhaustive
        row += 1
        wb.save("exhRes2.xlsx")


c_x = 100
c_y = 1.5
h1 = 1 # минимальная длина
h2 = 120 # максимальная длина
res = tuner(c_x, c_y, h1, h2)
