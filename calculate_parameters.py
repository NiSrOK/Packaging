import random
from genetic2D import perform_crossover_columns, divide_segments, get_best_columns
from openpyxl import Workbook, load_workbook
import plotly.graph_objects as go
import timeit
from classes2D import create_list_of_connections

def generate_base_population(n):
    list_of_segments = []
    for i in range(n):
        a = random.randint(1,10)
        b = random.randint(1,10)
        p = random.randint(1,20)
        name = i
        list_of_segments.append((a, b, p, name))
    return list_of_segments

def calculate_length(list_of_segments):
    length = 0
    for i in range(len(list_of_segments)):
        length += list_of_segments[i][0] + list_of_segments[i][1]
    return length

def population_preparation(list_of_segments, k, g, c_x, h1, h2, count_columns_base = 1):
    # Разделение отрезков на наборы соответствующие ограничениям по длине
    base_list_columns = divide_segments(list_of_segments, h1, h2)
    if len(base_list_columns) > count_columns_base:
        # Возьмем нужное количество столбцов, приоритет - максимальная длина
        best_list_columns = get_best_columns(base_list_columns, count_columns_base)
    else:
        # Если нужное количество столбцов уже соответствует - берем все 
        best_list_columns = base_list_columns
    # Получаем список списков с лучшими расстановками объектов внутри столбцов 
    connections = genetic_columns(best_list_columns[0], k, g, c_x, h1, h2)

    return connections

def genetic_columns(list_of_segments, k, g, c, h1, h2):
    list_of_columns = []
    # Разделение отрезков на наборы

    #создадим популяцию из объектов в одном столбце 
    population = create_list_of_connections(list_of_segments, k, c, h1, h2)

    for gen in range(1, g+1):
        # Рассчет пригодности для каждой перестановки
        fitness = []
        for perm in population:
            fitness.append((perm, perm.deviation))
        fitness.sort(key=lambda x: x[1])
        best_perm, best_dev = fitness[0]

        # Проверка на оптимальное решение
        if best_dev == 0:
            break

        # Генерация потомков
        num_children = k - 1
        children = []
        while num_children > 0:
            perm1, perm2 = random.sample(population, 2)
            child = perform_crossover_columns(perm1, perm2)
            child.calculate_connection()
            child.calculate_deviation(c, h1, h2)
            # Обмен двумя случайными сегментами в потомке
            if len(child.list_of_segments) > 1:
                i, j = random.sample(range(len(child.list_of_segments)), 2)
                child.list_of_segments[i], child.list_of_segments[j] = child.list_of_segments[j], child.list_of_segments[i]
            children.append(child)
            num_children -= 1

        # Объединение родительской и дочерней популяции
        combined_pop = fitness + [(child, child.deviation) for child in children]
        combined_pop.sort(key=lambda x: x[1])
        new_population = [combined_pop[i][0] for i in range(k)]
        
        population = new_population

    # Вывод наилучшей перестановки, ее значение пригодности, начального отступа
    find = False
    for res in fitness:
        if (res[0].a + res[0].b >= h1) and (res[0].a + res[0].b <= h2):
            return res[1]
            # find = True
            # list_of_columns.append(res[0])
            # break

    # if find == False:
    #     print('Не удалось найти расстановку удовлетворяющую заданным параметрам.')
    

    # return list_of_columns

def tuner(c, h1, h2):
    wb = Workbook()
    ws = wb.active
    ws[f'A1'] = 'k'
    ws[f'B1'] = 'g'
    ws[f'C1'] = 'Отклонение от целевого центра тяжести'
    ws[f'D1'] = 'Время расчета генетическим алгоритмом'
    row = 2
    glob_res = []
    population = generate_base_population(100)
    for k in range(2, 200, 5):
        for g in range(2, 200, 5):
            list_of_segments = population
            print(f'k = {k}, g = {g}')
            start_time_genetic = timeit.default_timer()
            gen_dev = population_preparation(list_of_segments, k, g, c, h1, h2)
            end_time_genetic = timeit.default_timer()
            total_time_genetic = end_time_genetic - start_time_genetic

            dev = float('{:.3f}'.format(gen_dev))
            print(dev)
            if dev == 0:
                dev = 0.001
            glob_res.append((int(k), int(g), dev))
            ws[f'A{row}'] = k
            ws[f'B{row}'] = g
            ws[f'C{row}'] = dev
            ws[f'D{row}'] = total_time_genetic
            # ws[f'E{row}'] = total_time_exhaustive / count
            row += 1
            wb.save("tuner1D100.xlsx")
    return glob_res

def plot_3d_graph(data):
    # Разделение данных на отдельные списки для каждой координаты
    x_data, y_data, z_data = zip(*data)

    # Создание трехмерного графика
    fig = go.Figure(data=[go.Scatter3d(x=x_data, y=y_data, z=z_data, mode='markers', marker=dict(color='red', size=3))])

    # Настройка меток осей и установка логарифмической шкалы для оси Z
    fig.update_layout(
        scene=dict(
            xaxis=dict(
                title=dict(text='k', font=dict(size=20)),  # Увеличение размера шрифта названия оси X
                tickfont=dict(size=14)  # Увеличение размера шрифта меток оси X
            ),
            yaxis=dict(
                title=dict(text='g', font=dict(size=20)),  # Увеличение размера шрифта названия оси Y
                tickfont=dict(size=14)  # Увеличение размера шрифта меток оси Y
            ),
            zaxis=dict(
                title=dict(text='Отклонение от целевого центра тяжести', font=dict(size=20)),  # Увеличение размера шрифта названия оси Z
                type='log',
                tickfont=dict(size=14)  # Увеличение размера шрифта меток оси Z
            ),
        )
    )

    # Отображение графика
    fig.show()

def read_data(file_path):
    # Открываем Excel-файл
    wb = load_workbook(file_path)
    ws = wb.active
    data = []
    # Проходим по строкам и считываем данные
    for row in range(2, ws.max_row + 1):
        # Получаем значения из столбцов A, B, и C для текущей строки
        a_value = int(ws[f'A{row}'].value)
        b_value = int(ws[f'B{row}'].value)
        c_value = float(ws[f'C{row}'].value)

        data.append((a_value, b_value, c_value))

    # Закрываем Excel-файл
    wb.close()

    return data


c = 50 # целевой центр тяжести
h1 = 1 # минимальная длина
h2 = 10000 # максимальная длина
# h1 = 1 # минимальная длина
# h2 = 500 # максимальная длина
# res = tuner(c, h1, h2)
res = read_data('tuner1D200_2.xlsx')
# data_points = [(1, 2, 3), (4, 5, 6), (7, 8, 9)]
plot_3d_graph(res)
