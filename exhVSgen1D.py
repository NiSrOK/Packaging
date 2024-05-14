import random
from genetic2D import genetic_columns
from openpyxl import Workbook, load_workbook
import timeit
from exhaustive_search import exhaustive_search

def get_best_columns(base_list_columns, count_columns):
    list_len_column = []

    for column in base_list_columns:
        len_column = 0
        for segment in column:
            len_column += segment[0] + segment[1]
        list_len_column.append((len_column, column))  # Сохраняем пару (длина столбца, столбец)

    # Сортируем столбцы по длине в убывающем порядке
    sorted_columns = sorted(list_len_column, reverse=True)
    print(sorted_columns)

    # Возвращаем первые count_columns самых длинных столбцов
    return [column for length, column in sorted_columns[:count_columns]]

# функция для разделения всех отрезков на наборы соответствующие ограничениям по длине
def divide_segments(segments, h1, h2):
    # segments.sort(key=lambda x: x[0])  # Сортируем отрезки по начальным точкам
    random.shuffle(segments)

    result_sets = []
    current_set = []
    current_length = 0
    not_in_any_set = []

    # segments_in_use = []
    # flag = True

    # while flag:
    #     if len(segments) == len(segments_in_use) or len(segments) == (len(segments_in_use) + len(not_in_any_set)):
    #         break
    for segment in segments:
        length = segment[0] + segment[1]  # Длина отрезка
        if length > h2:
            not_in_any_set.append(segment)
            continue
        if current_length + length <= h2:
            current_set.append(segment)
            current_length += length
        elif current_length >= h1:
            result_sets.append(current_set)
            current_set = [segment]
            current_length = length
        else:
            not_in_any_set.append(segment)

    # if current_length >= h1:
    #     result_sets.append(current_set)
    
    if current_length >= h1 and current_length <= h2:
        result_sets.append(current_set)
    elif current_length != 0:
        not_in_any_set.append(current_set)

    # Выводим отрезки, которые не попали ни в один набор
    if not_in_any_set:
        print("Отрезки, которые не попали ни в один набор:")
        for segment in not_in_any_set:
            print(segment)
    
    # result_connections = []
    # for set in result_sets:
    #     con = Connection(list_of_segments=set)
    #     result_connections.append(con)

    return result_sets

def population_preparation(list_of_segments, k, g, c_x, h1, h2, count_columns):
    # Разделение отрезков на наборы соответствующие ограничениям по длине
    base_list_columns = divide_segments(list_of_segments, h1, h2)
    if count_columns is not None and len(base_list_columns) > count_columns:
        # Возьмем нужное количество столбцов, приоритет - максимальная длина
        best_list_columns = get_best_columns(base_list_columns, count_columns)
    else:
        # Если нужное количество столбцов уже соответствует - берем все 
        best_list_columns = base_list_columns
    # Получаем список списков с лучшими расстановками объектов внутри столбцов 
    connections = genetic_columns(best_list_columns, k, g, c_x, h1, h2)

    return connections[0].deviation

def generate_base_population(n):
    list_of_segments = []
    for i in range(n):
        a = random.randint(1,20)
        b = random.randint(1,20)
        p = random.randint(1,20)
        name = i
        list_of_segments.append((a, b, p, name))
    return list_of_segments

def tuner(c_x, c_y, h1, h2):
    wb = Workbook()
    ws = wb.active
    ws[f'A1'] = 'Отклонение генетический 1 набор'
    ws[f'B1'] = 'Отклонение ПП лучший'
    ws[f'C1'] = 'Отклонение ПП худший'
    ws[f'D1'] = 'Время генетический 1 набор'
    ws[f'E1'] = 'Время ПП'
    row = 2
    for _ in range(20):
        population = generate_base_population(10)
        list_of_segments = population

        start_time_genetic = timeit.default_timer()
        gen_dev_total_1 = population_preparation(list_of_segments, 30, 30, c_x, h1, h2, count_columns=1)
        end_time_genetic = timeit.default_timer()
        total_time_genetic_1 = end_time_genetic - start_time_genetic
        print(f'Отклонение генетический {gen_dev_total_1}')

        
        start_time_exhaustive = timeit.default_timer()
        ex_dev_best, ex_dev_worst = exhaustive_search(list_of_segments, c_x, h1, h2)
        end_time_exhaustive = timeit.default_timer()
        total_time_exhaustive = end_time_exhaustive - start_time_exhaustive
        print(f'Отклонение ПП лучший {ex_dev_best}')
        print(f'Отклонение ПП худший {ex_dev_worst}')
        print('----------------')

        res_gen_1 = float('{:.3f}'.format(gen_dev_total_1))
        res_exh_best = float('{:.3f}'.format(ex_dev_best))
        res_exh_worst = float('{:.3f}'.format(ex_dev_worst))

        ws[f'A{row}'] = res_gen_1
        ws[f'B{row}'] = res_exh_best
        ws[f'C{row}'] = res_exh_worst
        ws[f'D{row}'] = total_time_genetic_1
        ws[f'E{row}'] = total_time_exhaustive
        row += 1
        wb.save("genVSex.xlsx")


c_x = 50
c_y = 1.5
h1 = 1 # минимальная длина
# h2 = 120 # максимальная длина
h2 = 500 # максимальная длина
res = tuner(c_x, c_y, h1, h2)
